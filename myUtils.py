#!/usr/bin/env python
# coding=utf-8
import datetime
import os
import shutil

import openpyxl as oxl

# 全局变量区域
from openpyxl.styles import Border, Side, Font, PatternFill

borderNumDic = {-1: None, 0: "thin"}


# 复制文件夹到指定位置,传入文件夹的绝对路径，以及复制位置及复制后的文件夹名
def copyFolderToPath(folderPath, aimPath):
    copyFlag = True
    # 检查文件夹是否存在
    if not os.path.exists(folderPath):
        copyFlag = False
    else:
        # 复制文件夹
        shutil.copytree(folderPath, aimPath)
        copyFlag = True
    return copyFlag


# 传入一个文件夹的绝对路径，删除这个文件夹
def deleteFolder(folderPath):
    try:
        shutil.rmtree(folderPath)
    except Exception as ex:
        print(ex)


# 传入一个文件的绝对路径，删除这个文件
def deleteFile(filePath):
    try:
        os.remove(filePath)
    except Exception as ex:
        print(ex)

# 获得精确到秒的当前时间
def getNowSeconed():
    formatStr = "%Y-%m-%d %H:%M:%S"
    nowDate = datetime.datetime.now()
    nowDateStr = nowDate.strftime(formatStr)
    return nowDateStr


# 获得excell的常用样式
def getExcellStyleDic():
    styleDic = {}

    # 单线边框
    thinBorder = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

    # 文字居中
    alignStyle = oxl.styles.Alignment(horizontal='center', vertical='center')
    leftStyle = oxl.styles.Alignment(horizontal='left', vertical='center')
    rightStyle = oxl.styles.Alignment(horizontal='right', vertical='center')

    # 加粗字体
    boldFont = Font(bold=True)
    hyperLinkFont = Font(color='0000FF')
    underLineFont = Font(underline='single')

    styleDic["thin"] = thinBorder
    styleDic["align"] = alignStyle
    styleDic["bold"] = boldFont
    styleDic["left"] = leftStyle
    styleDic["right"] = rightStyle
    styleDic["link"] = hyperLinkFont
    styleDic["underLine"] = underLineFont
    return styleDic


# 写入一个标准的excell表头（居中，单线框，加粗）
def writeExcellHead(ws, headArr):
    # 获得常用样式
    styleDic = getExcellStyleDic()
    # 写入表头
    for index, head in enumerate(headArr):
        ws.cell(row=1, column=index + 1).value = head
        ws.cell(row=1, column=index + 1).border = styleDic["thin"]
        ws.cell(row=1, column=index + 1).alignment = styleDic["align"]
        ws.cell(row=1, column=index + 1).font = styleDic["bold"]
    return ws


# 写入一个内容单元格
# borderNum表示该单元格的边框对象，其值可查询全局变量styleDic
# ifAlign是一个boolean对象，True表示居中
# hyperLink表示该单元格指向的链接，默认为None，表示不指向任何链接
# fgColor表示该单元格的背景颜色，为一个RGB16进制字符串，默认为“FFFFFF”（白色）
# otherAlign表示当ifAlign为False时指定的其他对齐方式，是一个数字型变量，默认为None，当其为0时表示左对齐，1为右对齐
def writeExcellCell(ws, row, column, value, borderNum, ifAlign, hyperLink=None, fgColor="FFFFFF", otherAlign=None):
    # 获得常用样式
    styleDic = getExcellStyleDic()
    # 获得指定单元格
    aimCell = ws.cell(row=row, column=column)
    # 设置值
    aimCell.value = value
    # 设置边框
    styleObjKey = borderNumDic[borderNum]
    if not styleObjKey:
        pass;
    else:
        styleObj = styleDic[styleObjKey]
        aimCell.border = styleObj
    # 设置居中
    if ifAlign:
        aimCell.alignment = styleDic["align"]
    elif otherAlign is not None:
        otherAlign = int(otherAlign)
        if otherAlign == 0:
            aimCell.alignment = styleDic["left"]
        else:
            aimCell.alignment = styleDic["right"]
    else:
        pass

    # 设置超链接
    if hyperLink:
        # 写入超链接
        aimCell.hyperlink = hyperLink
        # 设置当前单元格字体颜色为深蓝色，并添加下划线
        aimCell.font = styleDic["link"]
    else:
        pass

    # 设置填充颜色
    fill = PatternFill("solid", fgColor=fgColor)
    aimCell.fill = fill

    return ws


# 写入一个空格单元格，防止上一列文本超出
def writeExcellSpaceCell(ws, row, column):
    # 设置值
    ws.cell(row=row, column=column).value = " "

    return ws


# 设置excell的列宽
def setExcellColWidth(ws, colWidthArr):
    for colWidindex in range(len(colWidthArr)):
        ws.column_dimensions[chr(ord("A") + colWidindex)].width = colWidthArr[colWidindex]

    return ws


# 保存excell文件
def saveExcell(wb, saveName):
    savePath = ""
    # 处理传入的文件名
    saveName = saveName.split(".")[0] + ".xlsx"
    savePath = "{0}\\{1}".format(os.getcwd(), saveName)

    # 检测当前目录下是否有该文件，如果有则清除以前保存文件
    if os.path.exists(savePath):
        deleteFile(savePath)
    wb.save(savePath)
    return True
